# Constructor
class InternationalCaseObject:    
    PathFile = open('./Paths.txt')
    PVpath = PathFile.read()
    Text = None
    # Local Case Y/N
    LocalCase = None
    # Local Case Number
    LocalCaseNo = None
    # Global Case Number: 24b. MFR CONTROL NO.
    GlobalCaseNo = None        
    # Country: 1a. COUNTRY
    Country = None
    # Product Name: 14. SUSPECT DRUG(S) (include generic name)
    ProductName = None
    # Receipt Date: 24c. DATE RECEIVED
    ReceiptDate = None
    # Date Sent to Global PV: DATE OF THIS REPORT
    DateSentToGlobal = None
    # CIOMs Receipt date: Outlook
    # Serious Y/N: Find serious / non-serious
    Seriousness = None
    # Expected E/U: Find EXPECTEDNESS
    Expectedness = None
    # Reported to local authority Y/N/[NA] 
    # Reported on time  [empty]
    # Comments: Offlabel-use. / No need for reporting if benefit-risk-ratio has not changed.
    Comment = None
    
    def __init__(self):
        self.ReadPdf()
    # Pars
        # MFR-Control-No
        self.ReadPars("GlobalCaseNo","24b. MFR CONTROL NO.","25b. NAME AND ADDRESS","([A-Z]+-[A-Z]+-[0-9]+)") # Before, After, Pattern, Choices = None
        # MFR-Control-No
        self.ReadPars("Country","1a. COUNTRY","2. DATE OF BIRTH","([[A-Z][A-Z])$") # Before, After, Pattern, Choices = None
        # Receipt Date
        self.ReadPars("ReceiptDate","24c. DATE RECEIVED","DATE OF THIS REPORT","([0-9]+-\w+-[0-9]+)") # Before, After, Pattern, Choices = None
        # Date Sent to Global PV: DATE OF THIS REPORT
        self.ReadPars("DateSentToGlobal","DATE OF THIS REPORT","24d. REPORT SOURCE","([0-9]+-\w+-[0-9]+)") # Before, After, Pattern, Choices = None
        #ProductName
        self.ReadPars("ProductName","14. SUSPECT DRUG","15. DAILY DOSE","(#[0-9].*)") # Before, After, Pattern, Choices = None
        # Serious Y/N: Find serious / non-serious
        self.ReadPars("Seriousness","ADDITIONAL INFORMATION","TCPDF","(.*?)") # Before, After, Pattern, Choices = None
        # Expected E/U: Find EXPECTEDNESS
        self.ReadPars("Expectedness","ADDITIONAL INFORMATION","TCPDF","(.*?)") # Before, After, Pattern, Choices = None
        # Comments: Offlabel-use. / No need for reporting if benefit-risk-ratio has not changed.
        self.ReadPars("Comment","ADDITIONAL INFORMATION","TCPDF","(.*?)") # Before, After, Pattern, Choices = None
        self.FindLatestMail()
        self.WriteToXls()
        self.MoveToFolder()
        
    # Methods
    def ReadPdf(self):
        # Reads pdf
        import io
        import pdfminer
        from pdfminer.pdfinterp import PDFResourceManager, PDFPageInterpreter
        from pdfminer.converter import TextConverter
        from pdfminer.layout import LAParams
        from pdfminer.pdfpage import PDFPage
        
        path = self.PVpath
        rsrcmgr = PDFResourceManager()
        retstr = io.StringIO()
        codec = 'utf-8'
        laparams = LAParams()
        device = TextConverter(rsrcmgr, retstr, codec=codec, laparams=laparams)
        fp = open(path, 'rb')
        interpreter = PDFPageInterpreter(rsrcmgr, device)
        password = ""
        maxpages = 0
        caching = True
        pagenos = set()
    
        for page in PDFPage.get_pages(fp, pagenos, maxpages=maxpages,
                                      password=password,
                                      caching=caching,
                                      check_extractable=True):
            interpreter.process_page(page)
    
        self.Text = retstr.getvalue()
    
        fp.close()
        device.close()
        retstr.close()
        return self.Text
        
    def ReadPars(self, Par, Before, After, Pattern, Choices = None):
        import re
        # Read Pars
        Substring = re.search(Before + '(.*?)' + After,self.Text,re.DOTALL).group(1)  # extract coarse field
        Substring = Substring.replace("\r","") # get rid of linebreaks
        Substring = Substring.replace("\n","") # get rid of linebreaks
        ParString = re.search(Pattern, Substring,re.DOTALL).group(1) # search for pattern in field on CIOMS
        if Par == "Country":
            import pycountry
            CountryObjct = pycountry.countries.get(alpha_2=ParString)
            ParString = (CountryObjct.name)
        elif Par == "ProductName":
            product_list = []
            List_file = open('./Product_list.txt')
            product_list = List_file.read().splitlines()
            # create product dictionary: {unique search-term}{Name for File}
        elif Par == "Seriousness":
            print("Checking for seriousness...")
            if "non-serious" in Substring:
                print("non-serious")
                ParString = "N"
            elif "serious" in Substring:
                print("serious")
                ParString = "Y"
            else:
                print("nothing")
                ParString = None
        elif Par == "Expectedness":
            print("Checking for expectedness...")
            if "listed" in Substring:
                print("listed")
                ParString = "Y"
            elif "unlisted" in Substring:
                print("unlisted")
                ParString = "N"
            else:
                print("nothing")
                ParString = None
        elif Par == "Comment":
            print("Checking for Off-label use...")
            if "off-label" in Substring:
                print("Off-label use")
                ParString = "Offlabel-use. No need for reporting if benefit-risk-ratio has not changed."
            else:
                print("No off-label use")
                ParString = "No need for reporting if benefit-risk-ratio has not changed."
                
        exec("self." + Par + "= ParString")
        
    def FindLatestMail(self):
        print("Finding mail...")
        # Get Date from Outlook
    
    def WriteToXls(self):
        import openpyxl
        print("write stuf to xlsx file...")
        xfile = openpyxl.load_workbook("C:/Users/jonas/Dropbox/Monthly Recon/201801_CH_PV_Nuc.xlsx")
        sheet = xfile.get_sheet_by_name('Switzerland')
        row = sheet.max_row + 1
        sheet.cell(row = row, column = 1, value = "test")
        
        xfile.save("C:/Users/jonas/Dropbox/Monthly Recon/201801_CH_PV_Nuc.xlsx")
        print("...xslx filled out")
        # does sheet for actual month exist?
            # if not create it
            # if exist fill columns
                # NUC
                # Synacthen
        
    def CheckFolder(self): # check if folder exists, if not create one
       file_path = str('C:/Users/jonas/OneDrive/PV/CIOMS/Sorted CIOMS/' + folder)
       if not os.path.exists(file_path):
           os.makedirs(file_path)
           print('Created new folder for: ' + str(folder)  )
       return file_path
        
    def MoveToFolder(self):
        import shutil
        
        #shutil.move(src,dest)
        dest = "TEST"
        print("Moved to folder: " + dest)
        # drop in corresponding folder

    
    def convert_pdf_to_txt(path):
        import io
        import pdfminer
        from pdfminer.pdfinterp import PDFResourceManager, PDFPageInterpreter
        from pdfminer.converter import TextConverter
        from pdfminer.layout import LAParams
        from pdfminer.pdfpage import PDFPage
    
        rsrcmgr = PDFResourceManager()
        retstr = io.StringIO()
        codec = 'utf-8'
        laparams = LAParams()
        device = TextConverter(rsrcmgr, retstr, codec=codec, laparams=laparams)
        fp = open(path, 'rb')
        interpreter = PDFPageInterpreter(rsrcmgr, device)
        password = ""
        maxpages = 0
        caching = True
        pagenos = set()
    
        for page in PDFPage.get_pages(fp, pagenos, maxpages=maxpages,
                                      password=password,
                                      caching=caching,
                                      check_extractable=True):
            interpreter.process_page(page)
    
        text = retstr.getvalue()
    
        fp.close()
        device.close()
        retstr.close()
        return text


# import pycountry
x = InternationalCaseObject()
# print(x.Text)
print(x.Seriousness)
# print(x.ReceiptDate)
# print(x.DateSentToGlobal)
# print(type(x.ProductName))
print(x.ProductName)


